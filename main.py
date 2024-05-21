import xlrd
from openpyxl import Workbook, load_workbook
import time
import sys
import configparser
import msvcrt as msv

from openpyxl.styles import Font

start_time: time

# Default settings, columns for parse
SETTINGS = {
    'pass_col': 0,
    'ksg_col': 2,
    's_all_col': 29,
    'type_payment_col': 30,
    'department_col': 34,
    'koyko_dni_col': 38,
    'output_file_name': 'Output',
    'full_payment_str': 'Полная оплата',
}


def parsing_xlsx(file_path: str):
    print('Обработка докумета, подождите...')
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        rows = ws.rows
        data, departments = row_circle(rows)
        create_new_file(data, departments)
    except Exception as e:
        print(f'ошибка {e}')


def parsing_xls(file_path: str):
    print('Обработка докумета, подождите...')
    try:
        wb = xlrd.open_workbook(file_path)
        sh = wb.sheet_by_index(0)
        rows = sh.get_rows()
        data, departments = row_circle(rows)
        create_new_file(data, departments)
    except Exception as e:
        print(f'ошибка {e}')


def row_circle(rows):
    data = {}
    departments = []
    for row in rows:
        if row[0].value == 'C_I':
            pass
        elif row[SETTINGS['type_payment_col']].value == SETTINGS['full_payment_str']:
            department = row[SETTINGS['department_col']].value
            ksg = row[SETTINGS['ksg_col']].value
            s_all = string_to_float(row[SETTINGS['s_all_col']].value)
            koyko_dni = string_to_float(row[SETTINGS['koyko_dni_col']].value)
            if department not in departments:
                departments.append(department)
            if ksg != None:
                if ksg not in data.keys():
                    data[ksg] = {department: [s_all, 1, koyko_dni]}
                else:
                    if department not in data[ksg].keys():
                        data[ksg][department] = [s_all, 1, koyko_dni]
                    else:
                        data[ksg][department][0] += 1
                        data[ksg][department][0] += s_all
                        data[ksg][department][0] += koyko_dni
    return data, departments


def program_exit():
    print("---Прошло %.2f секунд ---" % (time.time() - start_time))
    print('Нажмите любую кнопку для выхода.')
    msv.getch()
    exit()


def string_to_float(string: str) -> float:
    try:
        return float(string)
    except:
        return 0.0


def create_new_file(data: dict, departments: list):
    wb = Workbook()
    ws = wb.active
    row = 8
    column_lambda = 2
    ws.cell(row=4, column=1, value='КСГ').font = Font(bold=True)
    ws.cell(row=7, column=1, value='Итого')
    ws.merge_cells(start_row=4, start_column=1, end_row=6, end_column=1)
    for department in departments:
        column_number = column_lambda + departments.index(department) * 3
        ws.cell(row=4, column=column_number, value=department)
        ws.merge_cells(start_row=4, start_column=column_number, end_row=4, end_column=column_number + 2)
        ws.cell(row=6, column=column_number, value='Пациентов').font = Font(bold=True)
        ws.cell(row=6, column=column_number + 1, value='Сумма').font = Font(bold=True)
        ws.cell(row=6, column=column_number + 2, value='К/дн').font = Font(bold=True)
        # ws.cell(row=7, column=column, value=f'=SUM({c_name_one}:{c_name_one})')
        # ws.cell(row=7, column=column + 1, value=f'=SUM()')
        # ws.cell(row=7, column=column + 2, value=f'=SUM()')

    for key, value in data.items():
        ws.cell(row=row, column=1, value=key)
        for department, values in value.items():
            column_number = column_lambda + departments.index(department) * 3
            ws.cell(row=row, column=column_number, value=values[0])
            ws.cell(row=row, column=column_number + 1, value=values[1])
            ws.cell(row=row, column=column_number + 2, value=values[2])
        row += 1
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        if column_letter != 'A':
            ws.cell(row=7, column=column[0].column, value=f'=SUM({column_letter}8:{column_letter}65535)')
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 0.7
        ws.column_dimensions[column_letter].width = adjusted_width
    wb.save(f'{SETTINGS["output_file_name"]}.xlsx')


def get_file_extension(file_path: str) -> str:
    try:
        return file_path.partition('.')[-1]
    except Exception as e:
        print(f'Произошла ошибка при попытке получить расширение файла.\n'
              f'{e}')
        program_exit()


def parsing_choice():
    try:
        if len(sys.argv) != 2:
            print(f'Похоже нет заагруженного файла или загружено несколько файлов.\n'
                  f' Для загрузки перенесите файл на ярлык программы.')
        elif get_file_extension(str(sys.argv[1])) == 'xlsx':
            parsing_xlsx(str(sys.argv[1]))
        elif get_file_extension(str(sys.argv[1])) == 'xls':
            parsing_xls(str(sys.argv[1]))
        else:
            print(f"Похоже загруженный файл неверного формата.\n Формат файла: {get_file_extension(sys.argv[1])}\n"
                  f" Допустимые форматы файла .xlsx и .xls\n")
    except Exception as e:
        print(f"Возникла ошибка при выборе режима работы программы.\n {e}")
    finally:
        program_exit()


def load_settings():
    config = configparser.ConfigParser()
    try:
        config.read('config.ini')
    except:
        print('Стандартные настройки.')
    else:
        # for key in config['strings']:
        #     print(config['strings'][key])
        for column in config['columns']:
            try:
                SETTINGS[column] = int(config['columns'][column])
            except:
                pass
        for string in config['strings']:
            try:
                SETTINGS[string] = config['string'][string]
            except:
                pass


if __name__ == '__main__':
    start_time = time.time()
    load_settings()
    parsing_choice()
